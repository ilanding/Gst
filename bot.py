import os
import re
import html
import base64
import shutil
import tempfile
from pathlib import Path

import qrcode
from openpyxl import load_workbook
from playwright.async_api import async_playwright
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, ContextTypes,
    MessageHandler, ConversationHandler, filters
)

BOT_TOKEN = os.getenv("BOT_TOKEN")
BASE_DIR = Path(__file__).parent
TEMPLATE = BASE_DIR / "template.html"

COLUMNS = [
    "Registration_Number", "Legal_Name", "Trade_Name", "Constitution",
    "Address", "Date_Liability", "Valid_From", "Valid_To",
    "Type_Registration", "Approving_Authority", "Authority_Name",
    "Authority_Designation", "Jurisdictional_Office", "Date_Issue"
]

WAIT_SHEET, WAIT_RANGE = range(2)
STATE = {}


def val(row, key):
    v = row.get(key, "")
    return "" if v is None else str(v).strip()


def safe_name(v):
    return re.sub(r"[^\w.-]+", "_", v)[:70] or "certificate"


def read_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    except StopIteration:
        wb.close()
        raise ValueError("Excel sheet empty hai.")

    missing = [x for x in COLUMNS if x not in headers]
    if missing:
        wb.close()
        raise ValueError(
            "Ye columns missing hain:\n" +
            "\n".join("• " + x for x in missing)
        )

    result = []
    for excel_row, values in enumerate(rows, start=2):
        if not any(v not in (None, "") for v in values):
            continue
        d = dict(zip(headers, values))
        d["_row"] = excel_row
        result.append(d)

    wb.close()
    return result


def parse_range(text, first_row, last_row):
    m = re.fullmatch(r"\s*(\d+)\s*[-:]\s*(\d+)\s*", text)
    if not m:
        raise ValueError("Format: 2-20")

    start, end = map(int, m.groups())

    if start < first_row or end < start or end > last_row:
        raise ValueError(f"Valid range {first_row}-{last_row} hai.")

    return start, end


def qr_payload(row):
    labels = {
        "Registration_Number": "Registration Number",
        "Legal_Name": "Legal Name",
        "Trade_Name": "Trade Name",
        "Constitution": "Constitution of Business",
        "Address": "Address",
        "Date_Liability": "Date of Liability",
        "Valid_From": "Valid From",
        "Valid_To": "Valid To",
        "Type_Registration": "Type of Registration",
        "Approving_Authority": "Approving Authority",
        "Authority_Name": "Authority Name",
        "Authority_Designation": "Authority Designation",
        "Jurisdictional_Office": "Jurisdictional Office",
        "Date_Issue": "Date of Issue",
    }

    return "\n".join(
        f"{labels[k]}: {val(row, k)}"
        for k in COLUMNS if val(row, k)
    )


def make_qr(row, path):
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=7,
        border=4
    )
    qr.add_data(qr_payload(row))
    qr.make(fit=True)
    qr.make_image().save(path)


def render(row, qr_path):
    text = TEMPLATE.read_text(encoding="utf-8")

    mapping = {
        "{{REGISTRATION_NUMBER}}": val(row, "Registration_Number"),
        "{{LEGAL_NAME}}": val(row, "Legal_Name"),
        "{{TRADE_NAME}}": val(row, "Trade_Name"),
        "{{CONSTITUTION}}": val(row, "Constitution"),
        "{{ADDRESS}}": val(row, "Address").replace("\n", "<br>"),
        "{{DATE_LIABILITY}}": val(row, "Date_Liability"),
        "{{VALID_FROM}}": val(row, "Valid_From"),
        "{{VALID_TO}}": val(row, "Valid_To"),
        "{{TYPE_REGISTRATION}}": val(row, "Type_Registration"),
        "{{APPROVING_AUTHORITY}}": val(row, "Approving_Authority"),
        "{{AUTHORITY_NAME}}": val(row, "Authority_Name"),
        "{{AUTHORITY_DESIGNATION}}": val(row, "Authority_Designation"),
        "{{JURISDICTIONAL_OFFICE}}": val(row, "Jurisdictional_Office"),
        "{{DATE_ISSUE}}": val(row, "Date_Issue"),
        "{{QR_DATA_URI}}": "data:image/png;base64," + base64.b64encode(qr_path.read_bytes()).decode("ascii"),
    }

    for key, value in mapping.items():
        if key == "{{QR_DATA_URI}}":
            text = text.replace(key, value)
        else:
            text = text.replace(key, html.escape(value))

    return text


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Welcome!\n\n"
        "📄 Bulk PDF Generator\n"
        "📱 Pehle .xlsx Excel/Google Sheets file upload karo."
    )
    return WAIT_SHEET


async def sheet_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document

    if not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text(
            "❌ Sirf .xlsx file upload karo.\n"
            "Google Sheets → File → Download → Microsoft Excel (.xlsx)"
        )
        return WAIT_SHEET

    user_id = update.effective_user.id
    workdir = Path(tempfile.mkdtemp(prefix=f"pdfbot_{user_id}_"))
    xlsx = workdir / "data.xlsx"

    try:
        tg_file = await doc.get_file()
        await tg_file.download_to_drive(custom_path=xlsx)
        rows = read_xlsx(xlsx)
    except Exception as e:
        shutil.rmtree(workdir, ignore_errors=True)
        await update.message.reply_text(f"❌ Error:\n{e}")
        return WAIT_SHEET

    if not rows:
        shutil.rmtree(workdir, ignore_errors=True)
        await update.message.reply_text("❌ Koi data row nahi mili.")
        return WAIT_SHEET

    STATE[user_id] = {"workdir": workdir, "rows": rows}

    await update.message.reply_text(
        f"✅ Sheet received.\n"
        f"Data rows: {rows[0]['_row']}–{rows[-1]['_row']}\n\n"
        "Ab range bhejo.\n"
        "Example: 2-20"
    )
    return WAIT_RANGE


async def range_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state = STATE.get(user_id)

    if not state:
        await update.message.reply_text("Pehle /start karke sheet upload karo.")
        return WAIT_SHEET

    rows = state["rows"]

    try:
        start, end = parse_range(
            update.message.text,
            rows[0]["_row"],
            rows[-1]["_row"]
        )
    except ValueError as e:
        await update.message.reply_text(f"❌ {e}")
        return WAIT_RANGE

    selected = [r for r in rows if start <= r["_row"] <= end]

    await update.message.reply_text(
        f"⚙️ {len(selected)} PDF generate ho rahe hain...\n"
        "Har PDF ka QR usi Excel row ke plain-text data se banega."
    )

    browser = None

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"]
            )

            page = await browser.new_page()
            qr_dir = state["workdir"] / "qr"
            pdf_dir = state["workdir"] / "pdf"
            qr_dir.mkdir()
            pdf_dir.mkdir()

            total = len(selected)

            for number, row in enumerate(selected, start=1):
                qr_file = qr_dir / f"qr_{row['_row']}.png"
                make_qr(row, qr_file)

                await page.set_content(
                    render(row, qr_file),
                    wait_until="networkidle"
                )

                pdf_file = pdf_dir / (
                    f"{row['_row']:04d}_"
                    f"{safe_name(val(row, 'Registration_Number'))}.pdf"
                )

                await page.pdf(
                    path=str(pdf_file),
                    format="A4",
                    print_background=True,
                    prefer_css_page_size=True,
                    margin={"top": "0", "right": "0", "bottom": "0", "left": "0"}
                )

                with pdf_file.open("rb") as f:
                    await update.message.reply_document(
                        document=f,
                        caption=(
                            f"📄 {number}/{total}\n"
                            f"Excel Row: {row['_row']}"
                        )
                    )

            await update.message.reply_text(
                f"✅ Done!\n{total} PDF Telegram par send ho gaye."
            )

    except Exception as e:
        await update.message.reply_text(
            f"❌ Generation error:\n{type(e).__name__}: {e}"
        )

    finally:
        if browser:
            try:
                await browser.close()
            except Exception:
                pass

        workdir = state.get("workdir")
        STATE.pop(user_id, None)

        if workdir:
            shutil.rmtree(workdir, ignore_errors=True)

    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state = STATE.pop(user_id, None)

    if state:
        shutil.rmtree(state["workdir"], ignore_errors=True)

    await update.message.reply_text("❌ Cancelled.")
    return ConversationHandler.END


def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN variable set nahi hai.")

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            WAIT_SHEET: [
                MessageHandler(filters.Document.ALL, sheet_handler)
            ],
            WAIT_RANGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, range_handler)
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )

    app.add_handler(conv)
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
